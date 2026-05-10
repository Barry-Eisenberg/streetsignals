"""Build a multi-tab Excel workbook for unmapped signal review.

Tabs:
- review_queue: data/unmapped_review_first_pass.csv
- quick_pick: concise initiative classification cue card
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
INPUT_CSV = DATA_DIR / "unmapped_review_first_pass.csv"
OUTPUT_XLSX = DATA_DIR / "unmapped_review_first_pass.xlsx"
INITIATIVE_TAXONOMY_PATH = DATA_DIR / "initiative-taxonomy.v1.json"


def autofit_columns(ws) -> None:
    max_lens = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            text = "" if value is None else str(value)
            max_lens[idx] = max(max_lens.get(idx, 0), len(text))

    for idx, length in max_lens.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(max(12, length + 2), 60)


def add_csv_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "review_queue"

    with INPUT_CSV.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.freeze_panes = "A2"
    autofit_columns(ws)


def add_quick_pick_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(title="quick_pick")

    rows = [
        ["Cue", "Recommended initiative classification(s)", "Notes"],
        ["Tokenized fund shares, treasuries, bonds, RWAs", "Tokenized Securities / RWA", "Assets on-chain"],
        ["Stablecoin, deposit token, on-chain cash", "Stablecoins & Deposit Tokens", "Digital money instrument"],
        ["Cross-border payment execution, payment rails", "Payment Infrastructure", "Flow execution and orchestration"],
        ["Clearing, settlement finality, custody, collateral, post-trade", "Settlement Infrastructure", "Market plumbing"],
        ["Enterprise blockchain platform, DLT network participation", "DLT / Blockchain Infrastructure", "Core infra build/participation"],
        ["Standards, interoperability, messaging bridges", "Interoperability & Standards", "Cross-system connectivity"],
        ["CBDC pilot/policy implementation", "CBDC", "Public money"],
        ["Protocol-native lending, AMM, decentralized derivatives", "DeFi", "Protocol finance"],
        ["Regulation, supervision, compliance obligations", "Regulatory / Compliance", "Policy and controls"],
        ["Enterprise roadmap, governance, strategic posture", "Digital Asset Strategy", "Operating model and governance"],
        ["Native crypto market activity (no institutional workflow)", "Crypto / Digital Assets", "Crypto-market specific"],
        ["", "", ""],
        ["Common multi-label combinations", "", ""],
        ["Tokenized issuance + post-trade servicing", "Tokenized Securities / RWA | Settlement Infrastructure", "Instrument + plumbing"],
        ["Stablecoin cash movement + rail execution", "Stablecoins & Deposit Tokens | Payment Infrastructure", "Instrument + payments"],
        ["DLT platform + connectivity standards", "DLT / Blockchain Infrastructure | Interoperability & Standards", "Core infra + standards"],
        ["Regulatory perimeter + stablecoin implementation", "Regulatory / Compliance | Stablecoins & Deposit Tokens", "Policy + implementation"],
        ["Strategy-driven platform build", "Digital Asset Strategy | DLT / Blockchain Infrastructure", "Roadmap + execution"],
        ["", "", ""],
        ["Theme projection", "", ""],
        ["tokenized theme", "Tokenized Securities / RWA", ""],
        ["stablecoins theme", "Stablecoins & Deposit Tokens | CBDC | Payment Infrastructure", ""],
        ["dlt theme", "DLT / Blockchain Infrastructure | Settlement Infrastructure | Interoperability & Standards | DeFi | Payment Infrastructure", ""],
        ["", "", ""],
        ["Tier -> Play index prior", "", "Override only with evidence; document in reviewer_notes"],
        ["Structural -> Play 3 (platform / partnership)", "tokenized-3 | stablecoins-3 | dlt-3", "Multi-party network or co-build framing"],
        ["Material -> Play 2 (expansion)", "tokenized-2 | stablecoins-2 | dlt-2", "Adding a second leg or client-facing rollout"],
        ["Context -> Play 1 (pilot)", "tokenized-1 | stablecoins-1 | dlt-1", "Single fund, counterparty, or workflow"],
        ["", "", ""],
        ["Cross-theme tie-breaker ladder", "", "Walk in order; stop at first decisive step"],
        ["1. Workflow specificity beats narrative", "", "Concrete workflow wins over generic noun"],
        ["2. Asset side vs. money side vs. rail", "tokenized | stablecoins | dlt", "Which leg of the trade does the signal describe?"],
        ["3. Institution type prior", "Asset Mgmt -> tokenized; Banks/Pmts -> stablecoins; FMI/CCP -> dlt", ""],
        ["4. Tier -> Play index", "", "See prior table above"],
        ["5. Persona alignment (audienceMatch)", "asset_managers | banks_fmis | fintech | policy_risk", ""],
        ["6. Source credibility + recency + prevalence", "", "From signal-strength-methodology.v1.json"],
        ["7. Adjudicator call", "", "Record tie_broken_by=adjudicator with one-line rationale"],
        ["", "", ""],
        ["Reason codes", "", ""],
        ["RC01_INSUFFICIENT_EVIDENCE", "keep_unmapped", "Text too vague to map confidently"],
        ["RC02_STRATEGY_ONLY", "keep_unmapped", "Strategic positioning without implementation"],
        ["RC03_NATIVE_CRYPTO_ONLY", "keep_unmapped", "Crypto market or venue activity, no institutional workflow"],
        ["RC04_TAXONOMY_GAP", "candidate_new_theme", "Valid signal, no current theme fits"],
        ["RC05_SOURCE_QUALITY", "keep_unmapped", "Source too weak or duplicative"],
        ["RC06_DUPLICATE_OR_NEAR_DUPLICATE", "keep_unmapped", "Same event already covered"],
        ["RC07_DATA_QUALITY", "keep_unmapped", "Bad institution/date/category fields"],
        ["RC08_MACRO_COMMENTARY", "keep_unmapped", "Macro/monetary policy; recommend re-tier out of Structural"],
        ["RC09_REGULATORY_PERIMETER", "candidate_new_theme", "Supervisory action shaping institutional perimeter"],
    ]

    for row in rows:
        ws.append(row)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Section header rows: "Common multi-label combinations", "Theme projection",
    # "Tier -> Play index prior", "Cross-theme tie-breaker ladder", "Reason codes".
    for row_idx in (14, 21, 26, 31, 41):
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = section_fill
            cell.font = Font(bold=True)

    ws.freeze_panes = "A2"
    autofit_columns(ws)


def add_play_catalog_sheet(wb: Workbook) -> None:
    """One row per Play so reviewers can copy the canonical play_id."""
    ws = wb.create_sheet(title="play_catalog")

    rows = [
        ["play_id", "theme", "index", "title", "audienceMatch (default play_audience first)"],
        ["tokenized-1", "tokenized", 1, "Quiet pilot anchored in an existing fund", "asset_managers"],
        ["tokenized-2", "tokenized", 2, "Tokenized cash + tokenized funds for treasury and collateral", "banks_fmis | fintech"],
        ["tokenized-3", "tokenized", 3, "Market infrastructure partnerships", "banks_fmis"],
        ["stablecoins-1", "stablecoins", 1, "Controlled treasury / B2B pilot", "banks_fmis | fintech"],
        ["stablecoins-2", "stablecoins", 2, "Client-facing settlement enhancement", "banks_fmis"],
        ["stablecoins-3", "stablecoins", 3, "Infrastructure partnership or network participation", "banks_fmis"],
        ["dlt-1", "dlt", 1, "Targeted post-trade / collateral use case", "banks_fmis"],
        ["dlt-2", "dlt", 2, "Tokenized assets + DLT post-trade integration", "asset_managers | banks_fmis"],
        ["dlt-3", "dlt", 3, "Strategic DLT platform participation or build", "banks_fmis"],
    ]

    for row in rows:
        ws.append(row)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.freeze_panes = "A2"
    autofit_columns(ws)


def add_initiative_options_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(title="initiative_options")

    taxonomy = json.loads(INITIATIVE_TAXONOMY_PATH.read_text(encoding="utf-8"))
    initiatives = [
        i
        for i in taxonomy.get("canonicalInitiatives", [])
        if i.get("active", True)
    ]

    # Keep one commonly used mapped label as a supplemental choice so the tab
    # reflects the working 12-option reviewer set.
    if len(initiatives) == 11:
        initiatives.append(
            {
                "name": "Cross-Border Payments",
                "group": "Payments",
                "description": "Alias mapped to Payment Infrastructure for matrix consistency.",
                "isMatrixCategory": True,
            }
        )

    ws.append(["Initiative Selection Options", "", "", ""])
    ws.append([
        "Initiative Classification",
        "Group",
        "Description",
        "Matrix Category",
    ])

    for item in initiatives:
        ws.append(
            [
                item.get("name", ""),
                item.get("group", ""),
                item.get("description", ""),
                "Yes" if item.get("isMatrixCategory") else "No",
            ]
        )

    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(bold=True, size=12)

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.freeze_panes = "A3"
    autofit_columns(ws)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build multi-tab workbook for unmapped review")
    parser.add_argument(
        "--output",
        default=str(OUTPUT_XLSX),
        help="Output workbook path",
    )
    args = parser.parse_args()

    if not INPUT_CSV.exists():
        raise FileNotFoundError(f"Missing input CSV: {INPUT_CSV}")

    output_path = Path(args.output)

    wb = Workbook()
    add_csv_sheet(wb)
    add_quick_pick_sheet(wb)
    add_play_catalog_sheet(wb)
    add_initiative_options_sheet(wb)
    wb.save(output_path)

    try:
        display_path = output_path.relative_to(ROOT)
    except ValueError:
        display_path = output_path
    print(f"Workbook created: {display_path}")


if __name__ == "__main__":
    main()
